import pandas as pd
from pathlib import Path
from datetime import datetime, date, timedelta
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from platform import system

# ==============================================================================
# FLAG CONFIGURATION
# Central source of truth for all processing flags.
#
# Properties:
#   - text: The user-facing message that appears in the Excel comment.
#   - crucial: (bool) If True, the cell will be highlighted red as it requires
#              manual intervention. If False, it's a non-critical notice.
# ==============================================================================
FLAG_CONFIG: dict[str, dict] = {
    'early_in': {
        'text': 'Info: Early Clock-In.',
        'crucial': False
    },
    'late_exit': {
        'text': 'Info: Late Clock-Out.',
        'crucial': False
    },
    'missing_first_in': {
        'text': 'CRITICAL: Missing First In.',
        'crucial': True
    },
    'missing_last_out': {
        'text': 'CRITICAL: Missing Last Out.',
        'crucial': True
    },
    'unpunched_breaks': {
        'text': 'Info: Breaktime inserted.',
        'crucial': True
    },
    'check_break_time': {
        'text': 'REVIEW: Breaktime.',
        'crucial': True
    },
    'for_manual_checking': {
        'text': 'REVIEW: Unresolved.',
        'crucial': True
    },
    # --- WARNING FLAGS (Crucial: True) ---
    'short_duration_day': {
        'text': 'REVIEW: Short workday.',
        'crucial': True
    },
    'long_duration_day': {
        'text': 'REVIEW: Long workday.',
        'crucial': True
    },
    'excessive_punches': {
        'text': 'REVIEW: Over-punched.',
        'crucial': True
    },
    'imbalanced_day': {
        'text': 'REVIEW: Punch Type.',
        'crucial': True
    },
}

def preprocess_sheet(df_raw: pd.DataFrame) -> pd.DataFrame | None:
    """
    Takes a raw DataFrame (from one sheet or file) and performs initial processing.
    1. Checks for and standardizes headers.
    2. Parses date/time columns.
    Returns a processed DataFrame or None if headers are invalid.
    """
    try:
        df: pd.DataFrame = process_headers(df_raw.copy())
    except ValueError:
        # This will be caught by the caller to log which sheet is skipped
        raise

    # Force date/time columns to string before processing
    df['DATE'] = df['DATE'].astype(str).str.strip()
    df['TIME'] = df['TIME'].astype(str).str.strip()
    datetime_str_series: pd.Series = df['DATE'] + ' ' + df['TIME']

    # Parse Datetime
    df['DATETIME'] = parse_dates(datetime_str_series, generate_date_formats(), day_first=True)

    # Identify and filter out rows where parsing failed
    if df['DATETIME'].isnull().any():
        raise TypeError(f'Found row(s) with unparseable date/time values from a sheet.')

    # If all rows were invalid, return an empty frame
    if df.empty:
        return pd.DataFrame()

    # Re-format DATE and TIME columns for the remaining valid rows
    df['DATE'] = df['DATETIME'].dt.strftime('%b %d')
    df['TIME'] = df['DATETIME'].dt.time
    return df

def read_input_file(file_path: str | Path) -> tuple[pd.DataFrame, list[str]]:
    """
    Reads a CSV or XLSX file, combines sheets if necessary, and returns a single DataFrame.

    -   For XLSX files, it reads all sheets. For each sheet, it checks if it
        contains the required headers. It then combines all sheets that have
        the required headers into a single DataFrame.
    -   For CSV files, it reads the file and checks for required headers.
    -   It returns a list of log messages about which sheets were skipped.
    -   It raises a ValueError if no valid data (with required headers) is found.

    Args:
        file_path: The path to the input file (.csv or .xlsx).

    Returns:
        A tuple containing:
        - A combined pandas DataFrame ready for processing.
        - A list of log messages generated during reading.

    Raises:
        ValueError: If the file format is unsupported or no sheets/file with
                    valid headers are found.
    """
    path: Path = Path(file_path)
    file_suffix: str = path.suffix.lower()
    logs: list = []

    if file_suffix == '.csv':
        df_raw = pd.read_csv(path)
        try:
            processed_df = preprocess_sheet(df_raw)
            if processed_df is None or processed_df.empty:
                raise ValueError(f'The CSV file {path.name=} has missing headers or no valid data.')
            logs.append(f'Successfully processed CSV file: {path.name=}')
            return processed_df, logs
        except ValueError as e:
            raise ValueError(f'The CSV file {path.name=} could not be processed. Details: {e}')

    elif file_suffix in ['.xlsx', '.xls']:
        try:
            xls_sheets = pd.read_excel(path, sheet_name=None, engine='openpyxl')
        except Exception as e:
            raise ValueError(f'Error reading Excel file {path.name=}: {e}')

        processed_dfs = []
        for sheet_name, sheet_df_raw in xls_sheets.items():
            if sheet_df_raw.empty:
                logs.append(f'Skipping empty sheet: {sheet_name=}.')
                continue
            try:
                processed_df = preprocess_sheet(sheet_df_raw)
                if not processed_df.empty:
                    processed_dfs.append(processed_df)
                    logs.append(f'Successfully processed sheet: {sheet_name=}.')
                else:
                    logs.append(f'Sheet {sheet_name=} contained no rows with valid data.')
            except ValueError:
                logs.append(f'Skipping sheet {sheet_name=} due to missing required headers.')
                continue

        if not processed_dfs:
            raise ValueError(f'No sheets with valid headers and data found in Excel file {path.name=}.')

        logs.append(f'Successfully combined {len(processed_dfs)} sheet(s) from {path.name}.')
        return pd.concat(processed_dfs, ignore_index=True), logs

    else:
        raise ValueError(f'Unsupported file format: {file_suffix=}. Please use a .csv or .xlsx file.')

def process_headers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Standardizes column headers and selects essential timesheet columns.

    This function performs the following steps:
    1.  Converts all column names to uppercase alphanumeric strings.
    2.  Searches for common aliases for 'ID', 'Name', 'Date', 'Time', and 'IN/OUT' columns.
    3.  If 'FIRSTNAME' and 'LASTNAME' are found, it combines them into a 'NAME' column.
    4.  Renames the identified columns to a standard set: 'ID', 'NAME', 'DATE', 'TIME', 'TYPE'.
    5.  Returns a new DataFrame containing only these standardized columns.

    Args:
        df: The raw input DataFrame from the timesheet file.

    Returns:
        A DataFrame with standardized and selected columns.

    Raises:
        ValueError: If any of the essential columns (ID, Name, Date, Time, Type)
                    cannot be found from the list of common aliases.
    """
    def standardize_col(col: str) -> str:
        return ''.join([char.upper() for char in col if char.isalnum()])
    df.columns = [standardize_col(col) for col in df.columns]

    standardized_columns: list[str] = list(df.columns)

    # Get ID column
    for col in ['PERSONNELID', 'PERSONNELNO', 'EMPLOYEENO', 'EMPLOYEEID', 'ID', 'EMPLOYEENUM',
                'PERSONNELNUM', 'EMPLOYEENUMBER', 'PERSONNELNUMBER', 'STAFFID', 'STAFFNO',
                'WORKERID', 'USERID', 'BADGENO', 'BADGEID', 'UID', 'RECORDID']:
        if col in standardized_columns:
            id_col = col
            break
    else:
        raise ValueError(f'PERSONNELID column not found in df with columns {standardized_columns}.')

    # Get Name column
    fname_col: str = ''
    for col in ['FIRSTNAME', 'GIVENNAME', 'FNAME', 'FORENAME', 'GNAME', 'PRIMERNOMBRE']:
        if col in standardized_columns:
            fname_col = col
            break
    lname_col: str = ''
    for col in ['LASTNAME', 'SURNAME', 'LNAME', 'FAMILYNAME', 'SECONDNAME', 'APELLIDO']:
        if col in standardized_columns:
            lname_col = col
            break
    if (fname_col != '') and (lname_col != ''):
        df['NAME'] = df[lname_col].str.strip().str.strip(',') + ', ' + df[fname_col].str.strip().str.strip(',')
        name_col = 'NAME'
    else:
        for col in ['NAME', 'FULLNAME', 'COMPLETENAME', 'EMPLOYEENAME',
                    'STAFFNAME', 'WORKERNAME', 'NOMBRE', 'NOM', 'USERNAME']:
            if col in standardized_columns:
                name_col = col
                break
        else:
            raise ValueError(f'NAME column not found in df with columns {standardized_columns}.')

    # Get date column
    for col in ['LOGDATE', 'DATE', 'DAY', 'WORKDATE', 'PUNCHDATE', 'TRANSDATE', 'ENTRYDATE', 'TDATE', 'FECHA', 'ARAW']:
        if col in standardized_columns:
            date_col = col
            break
    else:
        raise ValueError(f'DATE column not found in df with columns {standardized_columns}.')

    # Get time column
    for col in ['LOGTIME', 'LOGHOUR', 'HOUR', 'TIME', 'PUNCHTIME', 'ENTRYTIME',
                'TTIME', 'TRANSTIME', 'CLOCKTIME', 'HORA', 'ORAS']:
        if col in standardized_columns:
            time_col = col
            break
    else:
        raise ValueError(f'TIME column not found in df with columns {standardized_columns}.')

    # Get IN_OUT column
    for col in ['INOUT', 'LOGTYPE', 'TYPE', 'DIRECTION', 'ENTRYTYPE', 'STATUS', 'EVENT',
                'EVENTTYPE', 'INOUTTYPE', 'MOVEMENT', 'PUNCH', 'PUNCHTYPE', 'PUNCHINOUT', 'CATEGORY', 'CAT']:
        if col in standardized_columns:
            in_out_col = col
            break
    else:
        raise ValueError(f'IN & OUT column not found in df with columns {standardized_columns}.')

    # rename columns
    df = df.rename(columns={id_col: 'ID', name_col: 'NAME', date_col: 'DATE', time_col: 'TIME', in_out_col: 'TYPE'})
    return df[['ID', 'NAME', 'DATE', 'TIME', 'TYPE']]

def parse_dates(series: pd.Series, formats: list[str], day_first: bool = True) -> pd.Series:
    """
    Parses date strings using the specified hybrid strategy.

    1.  Fast Path: It iterates through `formats`. For each format, it
        attempts to parse the *entire* series. If successful, it returns
        the result immediately. This is highly efficient for uniform data.

    2.  Fallback Path: If no single format can parse the entire series, it
        falls back to using pandas' flexible but slower mixed-format parser.

    Args:
        series: The pandas Series of date strings.
        formats: A list of format strings to attempt for the fast path.
        day_first: A hint for the fallback parser for ambiguous dates (e.g., '01/02/2023').
                  Set to True for European (DMY), False for US (MDY).

    Returns:
        A pandas Series of datetime objects.
    """
    # Try to find one format that works for the whole series.
    for fmt in formats:
        try:
            # We use errors='raise'. If even one date fails, it will raise a
            # ValueError, and we'll move to the next format.
            parsed_series = pd.to_datetime(series, format=fmt, errors='raise')
            return parsed_series
        except (ValueError, TypeError):
            # This format didn't work for one or more dates.
            continue

    # --- FALLBACK PATH ---
    # If the loop finishes, no single format worked for the whole series.
    # Use format='mixed' to let pandas infer the format for each row individually.
    return pd.to_datetime(series, dayfirst=day_first, errors='coerce')

def generate_date_formats() -> list[str]:
    """
    Generates a list of date format strings in a specific, ordered sequence:
    1. ISO style (YYYY-MM-DD)
    2. European style (DD-MM-YYYY)
    3. US style (MM-DD-YYYY)

    Each style includes variations for:
    - Padded and non-padded day, month, and hour.
    - '/' and '-' separators.
    - Time with and without seconds.
    - 24-hour and 12-hour (AM/PM) clock.

    The order is preserved to ensure predictable parsing for ambiguous dates.
    """
    # 1. Define OS-specific modifiers for non-padded numbers
    # For day/month
    d_mod = '%-d' if system() != 'Windows' else '%#d'
    m_mod = '%-m' if system() != 'Windows' else '%#m'
    # For hour
    h_mod = '%-H' if system() != 'Windows' else '%#H'  # 24-hr
    i_mod = '%-I' if system() != 'Windows' else '%#I'  # 12-hr

    # 2. Define the building blocks for the format strings
    day_parts = ['%d', d_mod]
    month_parts = ['%m', m_mod]
    year_parts = ['%Y', '%y']
    separators = ['/', '-']

    # Comprehensive list of time formats to append
    time_formats = [
        # With seconds
        f'%H:%M:%S',  # 24h padded
        f'{h_mod}:%M:%S',  # 24h non-padded
        f'%I:%M:%S %p',  # 12h padded
        f'%I:%M:%S%p',
        f'{i_mod}:%M:%S %p',  # 12h non-padded
        f'{i_mod}:%M:%S%p',
        # Without seconds
        f'%H:%M',  # 24h padded
        f'{h_mod}:%M',  # 24h non-padded
        f'%I:%M %p',  # 12h padded
        f'%I:%M%p',
        f'{i_mod}:%M %p',  # 12h non-padded
        f'{i_mod}:%M%p',
    ]

    final_formats = []

    # 3. Generate formats in the specified order
    # --- ISO Formats (Y-M-D) ---
    iso_formats = []
    for sep in separators:
        for m in month_parts:
            for d in day_parts:
                for y in year_parts:
                    date_base = f'{y}{sep}{m}{sep}{d}'
                    for time in time_formats:
                        iso_formats.append(f'{date_base} {time}')
    final_formats.extend(iso_formats)

    # --- European Formats (D-M-Y) ---
    eur_formats = []
    for sep in separators:
        for d in day_parts:
            for m in month_parts:
                for y in year_parts:
                    date_base = f'{d}{sep}{m}{sep}{y}'
                    for time in time_formats:
                        eur_formats.append(f'{date_base} {time}')
    final_formats.extend(eur_formats)

    # --- US Formats (M-D-Y) ---
    us_formats = []
    for sep in separators:
        for m in month_parts:
            for d in day_parts:
                for y in year_parts:
                    date_base = f'{m}{sep}{d}{sep}{y}'
                    for time in time_formats:
                        us_formats.append(f'{date_base} {time}')
    final_formats.extend(us_formats)

    # 4. Remove duplicates while preserving order
    return list(dict.fromkeys(final_formats))

def create_grid(df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates a base grid for the final summary report.

    The grid has employee 'ID' and 'NAME' as initial columns. It then adds
    a column for each unique date present in the data, initialized to 0.0.

    Args:
        df: The processed DataFrame containing 'ID', 'NAME', and 'DATETIME' columns.

    Returns:
        A DataFrame structured as a grid for summarizing hours worked.
    """
    min_date: date = df['DATETIME'].min().date()
    max_date: date = df['DATETIME'].max().date()
    grid: pd.DataFrame = df.groupby('ID', as_index=False)['NAME'].first()

    # Add date columns
    for i in pd.date_range(min_date, max_date, freq='1d'):
        grid[f'{i:%b %d}']=0.0

    return grid.reset_index(drop=True)

def sort_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Sorts DataFrame, ensures consistent naming, and resolves name conflicts.

    This function performs three main tasks:
    1.  Establishes a canonical 'NAME' for each 'ID' based on the first
        occurrence to handle intra-employee name inconsistencies (e.g., 'Bob'/'Robert').
    2.  Identifies 'NAME's that are used by more than one unique 'ID' (a conflict).
    3.  For conflicting names, it creates a new unique name by appending the 'ID'
        (e.g., 'John Smith' becomes 'John Smith (101)').
    4.  Finally, it sorts the entire DataFrame by the resolved 'NAME' and then
        by 'DATETIME'.

    Args:
        df: The DataFrame to sort. It must contain 'ID', 'NAME', and 'DATETIME' columns.

    Returns:
        A sorted DataFrame with consistent and unique names for each employee,
        with the index reset.
    """
    # 1. Establish a canonical 'NAME' for each 'ID'
    # This handles cases where one person has multiple name entries (e.g., 'J. Doe', 'Jane Doe')
    id_to_name: pd.Series = df.drop_duplicates(subset=['ID']).set_index('ID')['NAME']
    df['NAME'] = df['ID'].map(id_to_name)

    # 2. Find names that are used by more than one ID
    # Group by name and count the number of unique IDs for each name
    name_counts = df.groupby('NAME')['ID'].nunique()
    # Filter to get only the names with more than one unique ID
    conflicting_names = name_counts[name_counts > 1].index

    # 3. Resolve conflicts by appending the ID to the name
    # Create a boolean mask for rows that have a conflicting name
    is_conflict = df['NAME'].isin(conflicting_names)

    # For the conflicting rows, update the NAME to 'NAME (ID)'
    # .loc is used to ensure we are modifying the DataFrame correctly
    df.loc[is_conflict, 'NAME'] = df.loc[is_conflict, 'NAME'] + ' (' + df.loc[is_conflict, 'ID'].astype(str) + ')'

    # 4. Sort the DataFrame by the resolved name and then by datetime
    df = df.sort_values(by=['NAME', 'DATETIME'], ascending=[True, True])

    return df.reset_index(drop=True)

def standardize_logtype(df: pd.DataFrame) -> pd.DataFrame:
    """
    Standardizes the punch type column to either 'in' or 'out'.

    It parses various common representations (e.g., 'C/In', 'c/out', 0, 1,
    boolean True/False) into a simple 'in' or 'out' string.

    Args:
        df: DataFrame with a 'TYPE' column to be standardized.

    Returns:
        The DataFrame with the 'TYPE' column standardized.

    Raises:
        TypeError: If an unparseable value is found in the 'TYPE' column.
        ValueError: If, after parsing, there are not exactly two unique types,
                    indicating a potential parsing failure.
    """
    # log type conversion
    def parse_logtype(x: str | bool) -> str:
        if isinstance(x, str):
            x = x.lower()
            if 'in' in x:
                return 'in'
            elif 'out' in x:
                return 'out'
            elif x == '0':
                return 'in'
            elif x == '1':
                return 'out'

        if isinstance(x, bool):
            if x:
                return 'out'
            else:
                return 'in'
        raise TypeError(f'Invalid LOGTYPE. Parsing {x} but failed.')

    df['TYPE'] = df['TYPE'].apply(parse_logtype)
    if df['TYPE'].nunique() != 2:
        raise ValueError('LOGTYPE was not completely parsed. Be sure IN/OUT column is properly formatted.')
    return df

def remove_same_type_duplicate(df: pd.DataFrame, buffer: timedelta) -> pd.DataFrame:
    """
    Removes duplicate or redundant punches within a specified time buffer.
    Only handles same type punches e.g. in-in or out-out

    This function implements complex logic to clean up punches that are too
    close together. It handles:
    - Duplicates of the same type (e.g., 'in', 'in').
    - It uses the 'VERIFIED' status to prioritize which punch to keep.

    Args:
        df: The sorted DataFrame with helper columns ('INDEX', 'VERIFIED').
        buffer: The timedelta within which punches are considered duplicates.

    Returns:
        A DataFrame with duplicate punches removed.
    """
    # Helper column
    df['PRIOR_INDEX'] = df['INDEX'].shift(1)

    to_be_removed: set = set()

    for employee_id in df['ID'].unique().tolist():
        filtered: pd.DataFrame = df[df['ID'] == employee_id]

        # if same_type and earlier entry is verified, remove later entry
        same_type_mask: pd.Series = (filtered['DATETIME'].diff() <= buffer) & (filtered['TYPE'] == filtered['TYPE'].shift(1))
        prior_verified: pd.Series = filtered.loc[same_type_mask & (filtered['VERIFIED'].shift(1) == True), 'INDEX']
        if not prior_verified.empty:
            to_be_removed.update(prior_verified)

        # if same_type and later entry is verified, remove earlier entry
        current_verified: pd.Series = filtered.loc[same_type_mask & (filtered['VERIFIED'] == True) &
                                                   (filtered['VERIFIED'].shift(1) == False), 'PRIOR_INDEX']
        if not current_verified.empty:
            to_be_removed.update(current_verified)

        # if same_type and unverified and type=='in', remove later entry
        indices: pd.Series = filtered.loc[same_type_mask &
                                          (filtered['VERIFIED'] == False) &
                                          (filtered['VERIFIED'].shift(1) == False) &
                                          (filtered['TYPE'] == 'in'), 'INDEX']
        if not indices.empty:
            to_be_removed.update(indices)

        # if same_type and unverified and type=='out', remove earlier entry
        indices = filtered.loc[same_type_mask &
                               (filtered['VERIFIED'] == False) &
                               (filtered['VERIFIED'].shift(1) == False) &
                               (filtered['TYPE'] == 'out'), 'PRIOR_INDEX']
        if not indices.empty:
            to_be_removed.update(indices)

    if to_be_removed:
        df = df[~df['INDEX'].isin(to_be_removed)]

    # Remove helper column
    df = df.drop('PRIOR_INDEX', axis='columns')
    return df

def remove_diff_type_duplicate(df: pd.DataFrame, buffer: timedelta) -> pd.DataFrame:
    """
    Removes redundant, close-proximity punches of different types (in-out or out-in).

    This function targets ambiguous punch pairs that occur within the specified time
    buffer. It's designed to clean up scenarios where a user might have
    accidentally punched twice in quick succession, creating an illogical sequence.

    It intelligently removes one of the punches in the pair by examining the
    surrounding verified punches to restore a logical pattern. For example:
    - In a sequence like `verified IN -> (IN-OUT) -> verified IN`, the pair is likely erroneous.
    - It handles `in-out` and `out-in` patterns by checking the punches that
      come two steps before and one step after the current position.

    Args:
        df: The sorted DataFrame with helper columns ('INDEX', 'VERIFIED').
        buffer: The timedelta within which punches are considered duplicates.

    Returns:
        A DataFrame with duplicate punches removed.
    """
    # Helper column
    df['PRIOR_INDEX'] = df['INDEX'].shift(1)

    to_be_removed: set = set()

    for employee_id in df['ID'].unique().tolist():
        filtered: pd.DataFrame = df[df['ID'] == employee_id]

        # in out -------------------------------------------------------------------------------------------------------
        in_out_mask: pd.Series = ((filtered['DATETIME'].diff() <= buffer) &
                                  (filtered['TYPE'].shift(1) == 'in') &
                                  (filtered['TYPE'] == 'out'))
        # prior
        in_in_out: pd.Series = filtered.loc[in_out_mask &
                                            (filtered['TYPE'].shift(2) == 'in') &
                                            (filtered['VERIFIED'].shift(2) == True), 'PRIOR_INDEX']
        if not in_in_out.empty:
            to_be_removed.update(in_in_out)

        # next
        in_out_out: pd.Series = filtered.loc[in_out_mask &
                                            (filtered['TYPE'].shift(-1) == 'out') &
                                            (filtered['VERIFIED'].shift(-1) == True), 'INDEX']
        if not in_out_out.empty:
            to_be_removed.update(in_out_out)

        # out in -------------------------------------------------------------------------------------------------------
        out_in_mask: pd.Series = ((filtered['DATETIME'].diff() <= buffer) &
                                  (filtered['TYPE'].shift(1) == 'out') &
                                  (filtered['TYPE'] == 'in'))
        # prior
        out_out_in: pd.Series = filtered.loc[out_in_mask &
                                            (filtered['TYPE'].shift(2) == 'out') &
                                            (filtered['VERIFIED'].shift(2) == True), 'PRIOR_INDEX']
        if not out_out_in.empty:
            to_be_removed.update(out_out_in)

        # next
        out_in_in: pd.Series = filtered.loc[out_in_mask &
                                            (filtered['TYPE'].shift(-1) == 'in') &
                                            (filtered['VERIFIED'].shift(-1) == True), 'INDEX']
        if not out_in_in.empty:
            to_be_removed.update(out_in_in)

    if to_be_removed:
        df = df[~df['INDEX'].isin(to_be_removed)]

    # Remove helper column
    df = df.drop('PRIOR_INDEX', axis='columns')
    return df

def adjust_first_in(df: pd.DataFrame, start_hour: timedelta, buffer: timedelta, flags: dict[str, list]) -> pd.DataFrame:
    """
    Adjusts the time of the first punch if it's before the official start time.

    - Flags punches that are significantly earlier than the start time ('early_in').
    - Modifies the timestamp of any punch before the start time (within a buffer)
      to be exactly at the start time.

    Args:
        df: The DataFrame to process.
        start_hour: The official start time of the workday as a timedelta.
        buffer: A timedelta defining the grace period around the start time.
        flags: A dictionary to which flags will be added. This is modified in place.

    Returns:
        The DataFrame with adjusted first-in times.
    """
    # --- Flagging Logic ---
    # Identify records to be flagged (time of day is 'way before' start_hour)
    to_be_flagged_mask: pd.Series = df['TIMEDELTA'] < start_hour - 2*buffer
    if not to_be_flagged_mask.empty:
        flag: list = df.loc[to_be_flagged_mask, ['ID', 'DATE']].to_dict(orient='records')
        flags.setdefault('early_in', []).extend(flag)

    # --- Time Modification Logic ---
    # Identify records for time modification (all records strictly before start_hour)
    to_be_modified_mask: pd.Series = df['TIMEDELTA'] < start_hour + buffer
    if not to_be_modified_mask.empty:
        df.loc[to_be_modified_mask, 'DATETIME'] = df['DATETIME'].dt.normalize() + start_hour

        # Reapply
        df['TIME'] = df['DATETIME'].dt.time
        df['TIMEDELTA'] = df['DATETIME'] - df['DATETIME'].dt.normalize()
    return df

def adjust_last_out(df: pd.DataFrame, end_hour: timedelta, buffer: timedelta, flags: dict[str, list]) -> pd.DataFrame:
    """
    Adjusts the time of the last punch if it's after the official end time.

    - Flags punches that are significantly later than the end time ('late_exit').
    - Modifies the timestamp of any punch after the end time (within a buffer)
      to be exactly at the end time.

    Args:
        df: The DataFrame to process.
        end_hour: The official end time of the workday as a timedelta.
        buffer: A timedelta defining the grace period around the end time.
        flags: A dictionary to which flags will be added. This is modified in place.

    Returns:
        The DataFrame with adjusted last-out times.
    """
    # --- Flagging Logic ---
    # Identify records to be flagged (time of day is 'way after' end_hour)
    to_be_flagged_mask: pd.Series = df['TIMEDELTA'] > end_hour + 2*buffer
    if not to_be_flagged_mask.empty:
        flag: list = df.loc[to_be_flagged_mask, ['ID', 'DATE']].to_dict(orient='records')
        flags.setdefault('late_exit', []).extend(flag)

    # --- Time Modification Logic ---
    # Identify records for time modification (all records strictly before start_hour)
    to_be_modified_mask: pd.Series = df['TIMEDELTA'] > end_hour - buffer
    if not to_be_modified_mask.empty:
        df.loc[to_be_modified_mask, 'DATETIME'] = df['DATETIME'].dt.normalize() + end_hour

        # Reapply
        df['TIME'] = df['DATETIME'].dt.time
        df['TIMEDELTA'] = df['DATETIME'] - df['DATETIME'].dt.normalize()
    return df

def verify_first_in(df: pd.DataFrame, break_time: dict[str, dict], morning_threshold: timedelta,
                    buffer: timedelta, flags: dict[str, list]) -> pd.DataFrame:
    """
    Verifies and corrects the first punch of the day.

    It identifies the first punch for each employee each day.
    - If it's correctly labeled 'in', it's marked as 'VERIFIED'.
    - If it's incorrectly labeled 'out' but occurs before a morning threshold
      or around a break time, it's corrected to 'in' and verified.
    - If an incorrect 'out' punch cannot be automatically corrected, it's
      flagged as 'missing_first_in'.

    Args:
        df: The DataFrame to process.
        break_time: A dictionary defining break periods.
        morning_threshold: A timedelta; punches before this are assumed to be 'in'.
        buffer: A grace period for time comparisons.
        flags: A dictionary for recording issues, modified in place.

    Returns:
        The DataFrame with first-in punches verified or corrected.
    """
    # get first entries of each day
    filtered: pd.DataFrame = df.groupby(['ID', 'DATE'], as_index=False, group_keys=False).first()

    # Tag verified for correctly labelled first_in
    indices: pd.Series = filtered.loc[filtered['TYPE'] == 'in', 'INDEX']
    if not indices.empty:
        mask: pd.Series = df['INDEX'].isin(indices)
        df.loc[mask, 'VERIFIED'] = True

    # Check incorrectly labelled first_in
    filtered = filtered[filtered['TYPE'] == 'out']
    if filtered.empty:
        return df

    # Filter for morning threshold
    indices = filtered.loc[filtered['TIMEDELTA'] <= morning_threshold + buffer, 'INDEX']
    if not indices.empty:
        mask = df['INDEX'].isin(indices)
        df.loc[mask, 'TYPE'] = 'in'
        df.loc[mask, 'VERIFIED'] = True

    # Filter for breaktime
    for breaktime_name, data in break_time.items():
        end: timedelta = data['end']

        indices = filtered.loc[(filtered['TIMEDELTA'] >= end-buffer) & (filtered['TIMEDELTA'] <= end+buffer), 'INDEX']
        if not indices.empty:
            mask = df['INDEX'].isin(indices)
            df.loc[mask, 'TYPE'] = 'in'
            df.loc[mask, 'VERIFIED'] = True

    # Flag missing first_in
    indices = df.loc[df['VERIFIED'] == False, 'INDEX']
    indices = filtered.loc[(filtered['TYPE'] == 'out') &
                           (filtered['INDEX'].isin(indices)), 'INDEX']
    if not indices.empty:
        mask = df['INDEX'].isin(indices)
        flag: list = df.loc[mask, ['ID', 'DATE']].to_dict(orient='records')
        flags.setdefault('missing_first_in', []).extend(flag)
    return df


def verify_last_out(df: pd.DataFrame, break_time: dict[str, dict], afternoon_threshold: timedelta,
                    buffer: timedelta, flags: dict[str, list]) -> pd.DataFrame:
    """
    Verifies and corrects the last punch of the day.

    It identifies the last punch for each employee each day.
    - If it's correctly labeled 'out', it's marked as 'VERIFIED'.
    - If it's incorrectly labeled 'in' but occurs after an afternoon threshold
      or around a break time, it's corrected to 'out' and verified.
    - If an incorrect 'in' punch cannot be automatically corrected, it's
      flagged as 'missing_last_out'.

    Args:
        df: The DataFrame to process.
        break_time: A dictionary defining break periods.
        afternoon_threshold: A timedelta; punches after this are assumed to be 'out'.
        buffer: A grace period for time comparisons.
        flags: A dictionary for recording issues, modified in place.

    Returns:
        The DataFrame with last-out punches verified or corrected.
    """
    # get last entries of each day
    filtered: pd.DataFrame = df.groupby(['ID', 'DATE'], as_index=False, group_keys=False).last()

    # Tag verified for correctly labelled last_out
    indices: pd.Series = filtered.loc[filtered['TYPE'] == 'out', 'INDEX']
    if not indices.empty:
        mask: pd.Series = df['INDEX'].isin(indices)
        df.loc[mask, 'VERIFIED'] = True

    # Check incorrectly labelled last_out
    filtered = filtered[filtered['TYPE'] == 'in']
    if filtered.empty:
        return df

    # Filter for afternoon threshold
    indices = filtered.loc[filtered['TIMEDELTA'] >= afternoon_threshold - buffer, 'INDEX']
    if not indices.empty:
        mask = df['INDEX'].isin(indices)
        df.loc[mask, 'TYPE'] = 'out'
        df.loc[mask, 'VERIFIED'] = True

    # Filter for breaktime
    for breaktime_name, data in break_time.items():
        start: timedelta = data['start']

        indices = filtered.loc[(filtered['TIMEDELTA'] >= start-buffer) & (filtered['TIMEDELTA'] <= start+buffer), 'INDEX']
        if not indices.empty:
            mask = df['INDEX'].isin(indices)
            df.loc[mask, 'TYPE'] = 'out'
            df.loc[mask, 'VERIFIED'] = True

    # Flag missing last_out
    indices = df.loc[df['VERIFIED'] == False, 'INDEX']
    indices = filtered.loc[(filtered['TYPE'] == 'in') &
                           (filtered['INDEX'].isin(indices)), 'INDEX']
    if not indices.empty:
        mask = df['INDEX'].isin(indices)
        flag: list = df.loc[mask, ['ID', 'DATE']].to_dict(orient='records')
        flags.setdefault('missing_last_out', []).extend(flag)
    return df

def add_helper_cols(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds temporary helper columns to the DataFrame for processing.

    - 'VERIFIED': A boolean flag, initialized False, to track processed punches.
    - 'INDEX': A copy of the DataFrame's index for stable referencing.
    - 'TIMEDELTA': The time portion of 'DATETIME' as a timedelta from midnight.

    Args:
        df: The DataFrame to which columns will be added.

    Returns:
        The DataFrame with new helper columns.
    """
    df['VERIFIED'] = False
    df['INDEX'] = df.index
    df['TIMEDELTA'] = df['DATETIME'] - df['DATETIME'].dt.normalize()
    return df

def adjust_break_time(df: pd.DataFrame, break_time: dict[str, dict], buffer: timedelta, flags: dict[str, list]) -> pd.DataFrame:
    """
    Adjusts punches around defined break times.

    For unpaid breaks, it snaps late clock-outs to the break start time and
    early clock-ins to the break end time.
    For paid breaks, it removes redundant in/out pairs that occur entirely
    within the break period.

    Args:
        df: The DataFrame to process.
        break_time: A dictionary defining break periods, start/end times, and paid status.
        buffer: A grace period for time comparisons.
        flags: A dictionary for recording issues (not used in this func but kept for signature consistency).

    Returns:
        The DataFrame with break time punches adjusted.
    """
    # Helper column
    df['NORMALIZED'] = df['DATETIME'].dt.normalize()
    df['PRIOR_INDEX'] = df['INDEX'].shift(1)
    df['NEXT_INDEX'] = df['INDEX'].shift(-1)
    to_be_flagged: set = set()

    for breaktime_name, data in break_time.items():
        start: timedelta = data['start']
        end: timedelta = data['end']

        verified_in: list[int] = []
        verified_out: list[int] = []

        # Auto-adjust Late Clock-outs to start of break time
        indices: pd.Series = df.loc[(df['TIMEDELTA'] >= start - buffer) &
                                    (df['TIMEDELTA'] < end - buffer) &
                                    (df['TYPE'] == 'out'), 'INDEX']
        if not indices.empty:
            mask: pd.Series = df['INDEX'].isin(indices)
            df.loc[mask, 'VERIFIED'] = True
            df.loc[mask, 'DATETIME'] = df.loc[mask, 'NORMALIZED'] + start

        # Auto-adjust Early Clock-ins to end of break time
        indices = df.loc[(df['TIMEDELTA'] <= end + buffer) &
                         (df['TIMEDELTA'] > start + buffer) &
                         (df['TYPE'] == 'in'), 'INDEX']
        if not indices.empty:
            mask = df['INDEX'].isin(indices)
            df.loc[mask, 'VERIFIED'] = True
            df.loc[mask, 'DATETIME'] = df.loc[mask, 'NORMALIZED'] + end

        # Auto-adjust Late Clock-outs near the end of breaktime.
        indices = df.loc[(df['TIMEDELTA'] >= end - buffer) &
                         (df['TIMEDELTA'] <= end + buffer) &
                         (df['TYPE'] == 'out'), 'INDEX']
        if not indices.empty:
            verified_in = []
            verified_out = []
            for i in indices:
                filtered: pd.DataFrame = df[df['INDEX'] == i]

                # check from prior
                prior_row: pd.DataFrame = df[df['INDEX'] ==  filtered['PRIOR_INDEX'].iat[0]]
                if not prior_row.empty:
                    prior_verified: bool = prior_row['VERIFIED'].iat[0]
                    prior_type: str = prior_row['TYPE'].iat[0]
                    prior_id: str = prior_row['ID'].iat[0]
                    if (prior_type == 'out') & (prior_verified == True) & (filtered['ID'].iat[0] == prior_id):
                        verified_in.append(i)
                        continue
                    elif (prior_type == 'in') & (prior_verified == True) & (filtered['ID'].iat[0] == prior_id):
                        verified_out.append(i)
                        continue

                # check from next
                next_row: pd.DataFrame = df[df['INDEX'] ==  filtered['NEXT_INDEX'].iat[0]]
                if not next_row.empty:
                    next_verified: bool = next_row['VERIFIED'].iat[0]
                    next_type: str = next_row['TYPE'].iat[0]
                    next_id: str = next_row['ID'].iat[0]
                    if (next_type == 'in') & (next_verified == True) & (filtered['ID'].iat[0] == next_id):
                        verified_out.append(i)
                        continue
                    elif (next_type == 'out') & (next_verified == True) & (filtered['ID'].iat[0] == next_id):
                        verified_in.append(i)
                        continue

                # cannot be verified
                to_be_flagged.add(i)

        # Auto-adjust Early Clock-ins near the start of breaktime.
        indices = df.loc[(df['TIMEDELTA'] >= start - buffer) &
                         (df['TIMEDELTA'] <= start + buffer) &
                         (df['TYPE'] == 'in'), 'INDEX']
        if not indices.empty:
            for i in indices:
                filtered = df[df['INDEX'] == i]

                # check from prior
                prior_row = df[df['INDEX'] == filtered['PRIOR_INDEX'].iat[0]]
                if not prior_row.empty:
                    prior_verified = prior_row['VERIFIED'].iat[0]
                    prior_type = prior_row['TYPE'].iat[0]
                    prior_id = prior_row['ID'].iat[0]
                    if (prior_type == 'out') & (prior_verified == True) & (filtered['ID'].iat[0] == prior_id):
                        verified_in.append(i)
                        continue
                    elif (prior_type == 'in') & (prior_verified == True) & (filtered['ID'].iat[0] == prior_id):
                        verified_out.append(i)
                        continue

                # check from next
                next_row = df[df['INDEX'] == filtered['NEXT_INDEX'].iat[0]]
                if not next_row.empty:
                    next_verified = next_row['VERIFIED'].iat[0]
                    next_type = next_row['TYPE'].iat[0]
                    next_id = next_row['ID'].iat[0]
                    if (next_type == 'in') & (next_verified == True) & (filtered['ID'].iat[0] == next_id):
                        verified_out.append(i)
                        continue
                    elif (next_type == 'out') & (next_verified == True) & (filtered['ID'].iat[0] == next_id):
                        verified_in.append(i)
                        continue

                # cannot be verified
                to_be_flagged.add(i)

        if len(verified_in) > 0:
            mask = df['INDEX'].isin(verified_in)
            df.loc[mask, 'TYPE'] = 'in'
            df.loc[mask, 'DATETIME'] = df.loc[mask, 'NORMALIZED'] + end
            df.loc[mask, 'VERIFIED'] = True

        if len(verified_out) > 0:
            mask = df['INDEX'].isin(verified_out)
            df.loc[mask, 'TYPE'] = 'out'
            df.loc[mask, 'DATETIME'] = df.loc[mask, 'NORMALIZED'] + start
            df.loc[mask, 'VERIFIED'] = True

    # Reapply
    df['PRIOR_INDEX'] = df['INDEX'].shift(1)
    df['TIMEDELTA'] = df['DATETIME'] - df['DATETIME'].dt.normalize()

    # Paid breaks
    for breaktime_name, data in break_time.items():
        start = data['start']
        end = data['end']
        paid: bool = data['paid']
        if not paid:
            continue

        # Remove out-in punches within paid breaks
        indices = df.loc[(df['TIMEDELTA'] <= end + buffer) & (df['TIMEDELTA'] >= start - buffer) &
                         (df['TYPE'] == 'in') &
                         (df['TIMEDELTA'].shift(1) >= start - buffer) & (df['TIMEDELTA'].shift(1) <= end + buffer) &
                         (df['TYPE'].shift(1) == 'out') &
                         (df['ID'] == df['ID'].shift(1)), ['INDEX', 'PRIOR_INDEX']]
        if not indices.empty:
            mask = df['INDEX'].isin(set(indices['INDEX'].tolist() + indices['PRIOR_INDEX'].tolist()))
            df = df[~mask]  # remove rows

        # Adjust clock in punches to the start
        indices = df.loc[(df['TIMEDELTA'] <= end + buffer) &
                         (df['TYPE'] == 'in') &
                         (df['TIMEDELTA'] >= start - buffer), 'INDEX']
        if not indices.empty:
            mask = df['INDEX'].isin(indices)
            df.loc[mask, 'DATETIME'] = df.loc[mask, 'NORMALIZED'] + start

    # Remove helper columns
    df = df.drop(['NORMALIZED', 'PRIOR_INDEX', 'NEXT_INDEX'], axis='columns')

    # Flag
    flag: list[dict] = df.loc[df['INDEX'].isin(to_be_flagged), ['ID', 'DATE']].to_dict(orient='records')
    flags.setdefault('check_break_time', []).extend(flag)
    return df

def verify_in_betweens(df: pd.DataFrame, flags: dict[str, list]) -> pd.DataFrame:
    """
    Attempts to fix single unverified punches located between two verified punches.

    This function iterates until no more fixes can be made. It handles:
    - An unverified punch between two 'in' punches -> corrected to 'out'.
    - An unverified punch between two 'out' punches -> corrected to 'in'.
    - Ambiguous cases (e.g., between an 'in' and an 'out') are removed and
      flagged for manual review.

    Args:
        df: DataFrame to process.
        flags: Dictionary for recording issues, modified in place.

    Returns:
        A DataFrame with in-between punches resolved.
    """
    indices_first_entry = df.groupby(['ID', 'DATE'], as_index=False, group_keys=False).first()['INDEX']
    indices_last_entry = df.groupby(['ID', 'DATE'], as_index=False, group_keys=False).last()['INDEX']
    flag: list = []

    # helper column
    df['PRIOR_TYPE'] = df['TYPE'].shift(1)
    df['NEXT_TYPE'] = df['TYPE'].shift(-1)

    prior_len: int = 0
    while True:
        filtered: pd.DataFrame = df[(df['VERIFIED'] == False) &
                                    (df['VERIFIED'].shift(1) == True) &
                                    (df['VERIFIED'].shift(-1) == True) &
                                    (~df['INDEX'].isin(indices_first_entry)) &
                                    (~df['INDEX'].isin(indices_last_entry))]
        if filtered.empty:
            break  # no more in-betweens
        elif len(filtered) == prior_len:
            break  # no more progress can be made
        else:
            prior_len = len(filtered)

        # In between two verified ins
        indices: pd.Series = filtered.loc[(filtered['PRIOR_TYPE'] == 'in') & (filtered['NEXT_TYPE'] == 'in'), 'INDEX']
        if not indices.empty:
            mask: pd.Series = df['INDEX'].isin(indices)
            df.loc[mask, 'TYPE'] = 'out'
            df.loc[mask, 'VERIFIED'] = True

        # In between two verified outs
        indices = filtered.loc[(filtered['PRIOR_TYPE'] == 'out') & (filtered['NEXT_TYPE'] == 'out'), 'INDEX']
        if not indices.empty:
            mask = df['INDEX'].isin(indices)
            df.loc[mask, 'TYPE'] = 'in'
            df.loc[mask, 'VERIFIED'] = True

        # Flag
        indices = filtered.loc[(filtered['PRIOR_TYPE'] == 'in') & (filtered['NEXT_TYPE'] == 'out'), 'INDEX']
        if not indices.empty:
            mask = df['INDEX'].isin(indices)
            flag.extend(df.loc[mask, ['ID', 'DATE']].to_dict(orient='records'))
            df = df[~mask]

        indices = filtered.loc[(filtered['PRIOR_TYPE'] == 'out') & (filtered['NEXT_TYPE'] == 'in'), 'INDEX']
        if not indices.empty:
            mask = df['INDEX'].isin(indices)
            flag.extend(df.loc[mask, ['ID', 'DATE']].to_dict(orient='records'))
            df = df[~mask]

    # Flag
    flags.setdefault('for_manual_checking', []).extend(flag)

    # Remove helper columns
    df = df.drop(['PRIOR_TYPE', 'NEXT_TYPE'], axis='columns')
    return df

def str_to_delta(time_str: str) -> timedelta:
    """
    Converts a time string (e.g., '07:00 AM') into a timedelta object.

    Args:
        time_str: The time string in 'HH:MM AM/PM' format.

    Returns:
        A timedelta object representing the time from midnight.
    """
    # time format is HH:MM AM/PM
    dt_obj = datetime.strptime(time_str, '%I:%M %p')
    return timedelta(hours=dt_obj.hour, minutes=dt_obj.minute)

def record_timestamps(df: pd.DataFrame, header: str | None = None) -> dict[tuple[str, str], str]:
    """
    Records all timestamps for each employee-day into a formatted string.

    This is used to create the 'before' and 'after' text for comments in the
    final Excel report.

    Args:
        df: The DataFrame containing the timestamps.
        header: An optional header string to prepend to each record (e.g., 'Raw Data:').

    Returns:
        A dictionary mapping (ID, DATE) tuples to a formatted string of their timestamps.
    """
    def format_group_entries(group_df_slice: pd.DataFrame) -> str:
        formatted_times: pd.Series = group_df_slice['TIME'].apply(lambda t: t.strftime('%I:%M %p'))
        strings_to_join: pd.Series = formatted_times + ' [' + group_df_slice['TYPE'] + ']\n'
        result: str = ''.join(strings_to_join)
        if header is None:
            return f'Raw Data:\n{result}'
        return f'{header}\n{result}'

    # Group by 'ID' and 'DATE', then apply the function
    result_series: pd.DataFrame = df.groupby(['ID', 'DATE']).apply(format_group_entries, include_groups=False)

    # Convert the resulting Series to a dictionary
    return result_series.to_dict()

def get_verifiable_sequence(df: pd.DataFrame) -> tuple[bool, list[str] | None]:
    """
    Determines if a sequence of punches can be logically ordered as alternating in/out.

    It checks if the existing 'VERIFIED' punches in a sequence contradict a simple
    alternating pattern. If no contradictions exist, it returns the valid
    alternating sequence.

    Args:
        df: A DataFrame subset for a single employee on a single day.

    Returns:
        A tuple containing:
        - bool: True if a valid sequence can be determined, False otherwise.
        - list[str] | None: The list of 'in'/'out' types for the sequence, or None if invalid.
    """
    if df.empty:
        return True, []

    possible_start_types: set = {'in', 'out'}

    def other_type(t) -> str:
        return 'out' if t == 'in' else 'in'

    for i, (_, row) in enumerate(df.iterrows()):
        if row['VERIFIED']:
            current_type: str = row['TYPE']

            if i % 2 == 0:
                invalid_start_if_this_pattern: str = other_type(current_type)
                possible_start_types.discard(invalid_start_if_this_pattern)
            else:
                invalid_start_if_this_pattern = current_type
                possible_start_types.discard(invalid_start_if_this_pattern)

            if not possible_start_types:
                return False, None

    if not possible_start_types:  # Should have been caught by the check inside the loop if n > 0
        return False, None

    if 'in' in possible_start_types:
        determined_start_type = 'in'
    elif 'out' in possible_start_types:  # This means 'in' was ruled out
        determined_start_type = 'out'
    else:
        # This state implies possible_start_types was non-empty but contained neither 'in' nor 'out'.
        # Given TYPE constraints, this indicates an error or unexpected input.
        return False, None

    final_sequence: list = []
    current_expected_type: str = determined_start_type

    for _ in range(len(df)):  # Iterate n times to build sequence of length n
        final_sequence.append(current_expected_type)
        current_expected_type = other_type(current_expected_type)

    return True, final_sequence

def pair_verification(df: pd.DataFrame, flags: dict[str, list]) -> pd.DataFrame:
    """
    Verifies punch sequences on a daily basis by assuming a simple in-out-out pattern.

    For each employee-day, if some punches are unverified, this function checks if
    the entire sequence can be made valid by forcing an alternating in/out pattern.
    If the existing verified punches don't contradict this pattern, it fills in
    the unverified ones. If there is a contradiction, the day is flagged for
    manual review.

    Args:
        df: The DataFrame to process.
        flags: Dictionary for recording issues, modified in place.

    Returns:
        The DataFrame with sequences verified.
    """
    # Helper column
    df['GROUP'] = df['ID'] + ' ' + df['DATE']

    to_be_flagged: list[int] = []
    for group in df['GROUP'].unique():
        filtered: pd.DataFrame = df[df['GROUP'] == group]
        if filtered[filtered['VERIFIED'] == False].empty:
            continue

        # check consecutive ins and outs
        verifiable, sequence = get_verifiable_sequence(filtered)
        if not verifiable:
            to_be_flagged.extend(filtered['INDEX'].tolist())
            continue

        indices = filtered['INDEX']
        mask = df['INDEX'].isin(indices)
        df.loc[mask, 'TYPE'] = sequence
        df.loc[mask, 'VERIFIED'] = True

    if len(to_be_flagged) > 0:
        flag: list[dict] = df.loc[df['INDEX'].isin(to_be_flagged), ['ID', 'DATE']].to_dict(orient='records')
        flags.setdefault('for_manual_checking', []).extend(flag)

    return df.drop('GROUP', axis='columns')

def flag_odd_groups(df: pd.DataFrame, flags: dict[str, list]) -> pd.DataFrame:
    """
    Flags any employee-day that has an odd number of punches.

    After all processing, a valid day should have an even number of punches
    (e.g., in/out, in/out/in/out). An odd number indicates a missing punch that
    could not be automatically fixed. It also flags any remaining unverified
    entries.

    Args:
        df: The final processed DataFrame.
        flags: Dictionary for recording issues, modified in place.

    Returns:
        The DataFrame, passed through without modification.
    """
    # Helper column
    df['GROUP'] = df['ID'] + ' ' + df['DATE']

    to_be_flagged: list[int] = []
    for group in df['GROUP'].unique():
        filtered: pd.DataFrame = df[df['GROUP'] == group]
        if len(filtered) % 2 != 0:
            to_be_flagged.extend(filtered['INDEX'].tolist())

    if len(to_be_flagged) > 0:
        flag: list[dict] = df.loc[df['INDEX'].isin(to_be_flagged), ['ID', 'DATE']].to_dict(orient='records')
        flags.setdefault('for_manual_checking', []).extend(flag)

    # Flag remaining unverified entries
    flag = df.loc[df['VERIFIED'] == False, ['ID', 'DATE']].to_dict(orient='records')
    if len(flag) > 0:
        flags.setdefault('for_manual_checking', []).extend(flag)

    return df.drop('GROUP', axis='columns')

def adjust_and_round(df: pd.DataFrame, round_to: list[timedelta], buffer: timedelta) -> pd.DataFrame:
    """
    Rounds punch times to the nearest specified time if they fall within a buffer.

    For example, if `round_to` includes 4:00 PM and buffer is 15 mins, any punch
    between 3:45 PM and 4:15 PM will be changed to exactly 4:00 PM.

    Args:
        df: The DataFrame to process.
        round_to: A list of timedelta objects to round to.
        buffer: The timedelta buffer around the rounding times.

    Returns:
        The DataFrame with times rounded.
    """
    def round_time(row: pd.Series) -> datetime:
        current_time: timedelta = row['TIMEDELTA']
        for time in round_to:
            if (current_time >= time - buffer) & (current_time <= time + buffer):
                return row['DATETIME'].normalize() + time
        return row['DATETIME']

    df['DATETIME'] = df.copy().apply(round_time, axis='columns')

    # Reapply
    df['TIME'] = df['DATETIME'].dt.time
    df['TIMEDELTA'] = df['DATETIME'] - df['DATETIME'].dt.normalize()
    return df

def remove_duplicate_dicts(dict_list: list[dict]) -> list[dict]:
    """
    Removes duplicate dictionaries from a list.

    Args:
        dict_list: A list that may contain duplicate dictionaries.

    Returns:
        A list of unique dictionaries.
    """
    seen: set = set()
    unique_dicts: list[dict] = []
    for d in dict_list:
        # Convert dict to a sorted tuple of items, so order doesn't matter
        items: tuple = tuple(sorted(d.items()))
        if items not in seen:
            seen.add(items)
            unique_dicts.append(d)
    return unique_dicts

def create_summary(df: pd.DataFrame):
    """
    Calculates the total hours worked per employee per day and creates a summary grid.

    This function uses an efficient pivot table approach to calculate the duration
    between each 'in' and 'out' pair and sum them up for each day.

    Args:
        df: The fully processed DataFrame with alternating 'in'/'out' punches.

    Returns:
        A summary DataFrame with employees as rows, dates as columns, and total
        worked hours as values.
    """
    summary: pd.DataFrame = create_grid(df)
    summary = summary.set_index('ID')

    # Helper column
    df['PREV_TYPE'] = df['TYPE'].shift(1)
    df['PREV_DATETIME'] = df['DATETIME'].shift(1)
    date_columns: list[str] = list(summary.columns.difference(['ID', 'NAME']))

    for employee_id in df['ID'].unique():
        for day in date_columns:
            filtered: pd.DataFrame = df[(df['ID'] == employee_id) &
                                        (df['DATE'] == day) &
                                        (df['TYPE'] == 'out') &
                                        (df['PREV_TYPE'] == 'in')]
            if filtered.empty:
                continue
            hours_worked: timedelta = (filtered['DATETIME'] - filtered['PREV_DATETIME']).sum()
            summary.at[employee_id, day] = hours_worked.total_seconds()/3600.0

    df.drop(['PREV_TYPE', 'PREV_DATETIME'], axis='columns', inplace=True)
    return summary.reset_index(drop=False)

def combine_comments(timestamp_old: dict[tuple[str, str], str],
                     timestamp_new: dict[tuple[str, str], str],
                     flags: dict[str, list],
                     silence_flags: list[str] | None = None) -> dict[tuple[str, str], str]:
    """
    Combines raw timestamps, processed timestamps, and flags into comment strings.

    Args:
        timestamp_old: Dict of raw timestamp strings.
        timestamp_new: Dict of processed timestamp strings.
        flags: Dict of all generated flags.
        silence_flags: A list of flag keys (from FLAG_CONFIG) to ignore when generating comments.

    Returns:
        A dictionary mapping (ID, DATE) tuples to a final, combined comment string.
    """
    if silence_flags is None:
        silence_flags = []

    comments: dict[tuple[str, str], str] = {}
    all_keys = set(timestamp_old.keys()) | set(timestamp_new.keys())

    # De-duplicate flags first
    for key, values in flags.items():
        flags[key] = remove_duplicate_dicts(values)

    # Add flags to any relevant keys
    for flag_name, flagged_items in flags.items():
        if flag_name in silence_flags:
            continue
        config = FLAG_CONFIG.get(flag_name)
        if not config:
            print(f'Warning: No config for flag {flag_name}.')
            continue

        comment_to_add = config['text']
        for item in flagged_items:
            key = (item['ID'], item['DATE'])
            all_keys.add(key)
            if key in comments:
                if comment_to_add not in comments[key]:
                    comments[key] += f'\n{comment_to_add}'
            else:
                comments[key] = comment_to_add

    # Now, append the timestamp data to every key that has content
    for key in all_keys:
        flag_part = comments.get(key, '')
        old_ts_part = timestamp_old.get(key, '')
        new_ts_part = timestamp_new.get(key, 'Processed Data:\n[No punches after processing]')

        # Combine all parts with clear separators
        final_comment_parts = []
        if flag_part:
            final_comment_parts.append(flag_part + '\n')
        if old_ts_part:
            final_comment_parts.append(old_ts_part)

        final_comment_parts.append(new_ts_part)
        final_comment_parts.append('✨[rs_uy]')

        comments[key] = '\n'.join(final_comment_parts)

    return comments

def insert_breaks(df: pd.DataFrame, break_times: dict[str, dict],
                  flags: dict[str, list]) -> pd.DataFrame:
    """
    Inserts missing break punches for unpaid breaks.

    If an employee has an 'in' punch before an unpaid break and an 'out' punch
    after it without any punches in between, this function inserts a corresponding
    'out' at the break start and an 'in' at the break end.

    Args:
        df: The DataFrame to process.
        break_times: Dictionary defining break periods.
        flags: Dictionary for recording issues, modified in place.

    Returns:
        A new DataFrame with missing break punches inserted and re-sorted.
    """
    flag: list[dict] = []
    new_rows_data: list[dict] = []

    # Group by ID and Date for daily processing
    grouped_df = df.groupby(['ID', 'DATE'])

    for (employee_id, row_date), group in grouped_df:
        group = group.sort_values(by='DATETIME')
        for break_name, break_data in break_times.items():
            if break_data['paid']:
                continue  # Only process unpaid breaks

            break_start: timedelta = break_data['start']
            break_end: timedelta = break_data['end']

            last_in_row: pd.Series | None = None
            for idx, row in group.iterrows():
                if row['TYPE'] == 'in':
                    last_in_row = row
                elif row['TYPE'] == 'out' and last_in_row is not None:
                    # Check if this in-out pair spans the unpunched break
                    if (last_in_row['TIMEDELTA'] < break_start) and (row['TIMEDELTA'] > break_end):

                        # Found a spanning segment without explicit break punches
                        # Add new 'out' punch at break start
                        break_start_datetime: datetime = group.iloc[0]['DATETIME'].normalize() + break_start
                        break_start_time = break_start_datetime.time()
                        break_end_datetime: datetime = group.iloc[0]['DATETIME'].normalize() + break_end
                        break_end_time = break_end_datetime.time()

                        new_rows_data.append({
                            'ID': employee_id,
                            'NAME': group.iloc[0]['NAME'],  # Get name from any row in group
                            'DATE': row_date,
                            'TIME': break_start_time,
                            'DATETIME': break_start_datetime,
                            'TYPE': 'out',
                            'TIMEDELTA': break_start,
                            'VERIFIED': True  # These are programmatically inserted, so 'verified'
                        })
                        # Add new 'in' punch at break end
                        new_rows_data.append({
                            'ID': employee_id,
                            'NAME': group.iloc[0]['NAME'],  # Get name from any row in group
                            'DATE': row_date,
                            'TIME': break_end_time,
                            'DATETIME': break_end_datetime,
                            'TYPE': 'in',
                            'TIMEDELTA': break_end,
                            'VERIFIED': True
                        })

                        # Flag this for comments
                        flag.append({'ID': employee_id, 'DATE': row_date})
                        # Move to next group, as this one has been handled
                        break
                    last_in_row = None  # Reset after processing an out punch

    # Flag
    if len(flag) > 0:
        flags.setdefault('unpunched_breaks', []).extend(flag)

    if new_rows_data:
        new_rows: pd.DataFrame = pd.DataFrame(new_rows_data)
        # Assign temporary unique index to new rows before concat. High numbers to not clash with existing indices.
        # This will be overwritten by reset_index anyway, but ensures temp validity.
        new_rows['INDEX'] = range(df['INDEX'].max() + 1, df['INDEX'].max() + 1 + len(new_rows))

        # Concatenate original df with new rows
        df_combined: pd.DataFrame = pd.concat([df, new_rows], ignore_index=True)

        # Re-sort the entire DataFrame by NAME, DATETIME
        df_combined = sort_df(df_combined)  # sort_df also re-calculates NAME column based on ID

        # Re-index the combined DataFrame to ensure 'INDEX' column is continuous and correct
        df_combined = df_combined.reset_index(drop=True)
        df_combined['INDEX'] = df_combined.index

        # Re-calculate TIMEDELTA to ensure it's correct for new rows (and existing ones if dates changed)
        df_combined['TIMEDELTA'] = df_combined['DATETIME'] - df_combined['DATETIME'].dt.normalize()
        df_combined['TIME'] = df_combined['DATETIME'].dt.time  # Also update TIME from DATETIME

        return df_combined
    return df

def get_crucial_flags(flags: dict[str, list]) -> list[tuple[str, str]]:
    """
    Extracts a de-duplicated list of cells that require crucial manual attention.

    Cruciality is determined by the `crucial: True` property in the global
    FLAG_CONFIG dictionary.

    Args:
        flags: The dictionary containing all generated flags.

    Returns:
        A de-duplicated list of (ID, DATE) tuples for crucial flags.
    """
    crucial_cells: set[tuple[str, str]] = set()

    # Iterate through the flags that were actually raised for this run
    for flag_name, flagged_items in flags.items():
        config: dict = FLAG_CONFIG.get(flag_name)
        # Check if the flag is defined in our config and is marked as crucial
        if config and config['crucial']:
            # Add all cells associated with this crucial flag to our set
            for item in flagged_items:
                crucial_cells.add((item['ID'], item['DATE']))

    return list(crucial_cells)

def find_writable_filename(output_path: str) -> Path:
    """
    Checks if a file is writable. If not, finds a unique alternative.

    If the target file is locked (e.g., open in Excel), it will append a
    counter to the filename (e.g., 'file (1).xlsx') until it finds an
    available path.

    Args:
        output_path: The desired output file path.

    Returns:
        A Path object for a writable file path.

    Raises:
        Exception: If an unexpected error other than PermissionError occurs.
    """
    path: Path = Path(output_path)
    try:
        # Attempt to open the file in append mode to check for write permissions.
        # This will raise PermissionError if the file is open/locked.
        with open(path, 'a'):
            pass
        # If successful, the original path is writable.
        return path
    except PermissionError:
        print(f'Warning: {path=} is currently open or you dont have write permissions.')
        print('Attempting to save with a new name...')

        counter: int = 1
        while True:
            # Create a new filename like 'timesheet_summary (1).xlsx'
            new_path: Path = path.with_name(f'{path.stem} ({counter}){path.suffix}')
            try:
                with open(new_path, 'a'):
                    pass
                print(f'File will be saved as {new_path=}')
                return new_path
            except PermissionError:
                # This new name is also locked, try the next one
                counter += 1
            except Exception as e:
                # Handle other potential file system errors
                print(f'An unexpected error occurred trying to find a writable filename: {e}')
                raise  # Re-raise the exception as it's not a simple permission issue
    except Exception as e:
        print(f'An unexpected error occurred checking file permissions: {e}')
        raise

def create_sheet(df: pd.DataFrame,
                 comments: dict[tuple[str, str], str],
                 flags: list[tuple[str, str]],
                 output_filename: str = 'timesheet_summary.xlsx') -> Path | None:
    """
    Creates a formatted Excel sheet from a summary DataFrame.

    This function writes the summary data to an Excel file and applies rich
    formatting, including:
    - Header styles (color, font).
    - Cell borders and alignment.
    - Conditional formatting (a red flag) for cells requiring manual review.
    - Detailed comments on each data cell showing raw and processed timestamps.
    - Auto-adjusted column widths.
    - Frozen panes for easy scrolling.

    Args:
        df: The input summary DataFrame.
        comments: A dictionary mapping (ID, date_column) to a comment string.
        flags: A list of (ID, date_column) tuples to be flagged with a red background.
        output_filename: The name of the output Excel file.
    """
    # Find a writable path before proceeding
    try:
        final_output_path: Path = find_writable_filename(output_filename)
    except Exception as e:
        print(f'Could not secure a writable output file. Aborting. Error: {e}')
        return  None

    # This makes the code easier to read and maintain.
    header_fill = PatternFill(start_color='3A3838', end_color='3A3838', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    flag_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    flag_font = Font(color='9C0006')
    id_name_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    thin_border_side = Side(style='thin', color='A6A6A6')
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    center_align = Alignment(horizontal='center', vertical='center')

    # Column names can also be constants
    id_col, name_col = 'ID', 'NAME'

    with pd.ExcelWriter(final_output_path, engine='openpyxl') as writer:
        # 1. Write the DataFrame to an Excel sheet
        sheet_name = 'Timesheet'
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 2. Get the workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # 3. Create efficient lookups for row/col indices and column names
        id_to_row = {employee_id: i + 2 for i, employee_id in enumerate(df[id_col])}
        date_to_col = {col_name: i + 1 for i, col_name in enumerate(df.columns)}

        # 4. Consolidated Formatting Loop (Single Pass)
        # This single loop applies all base formatting to the header and data cells.
        for row_idx in range(1, worksheet.max_row + 1):
            is_header = (row_idx == 1)
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # Apply formatting common to all cells
                cell.border = cell_border

                if is_header:
                    # Header-specific formatting
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                else:
                    # Data cell formatting
                    if col_idx <= 2:  # ID and NAME columns
                        cell.fill = id_name_fill
                        # Left-align NAME, center-align ID
                        if col_idx == date_to_col[name_col]:
                            cell.alignment = Alignment(vertical='center') # Default horizontal is left
                        else:
                            cell.alignment = center_align
                    else: # Date columns
                        cell.number_format = '0.00'
                        cell.alignment = center_align

        # 5. Apply Red Flags (Overrides base formatting)
        # This is kept separate as it's more efficient to iterate over the
        # small `flags` list than to check every cell for a flag.
        for employee_id, date_col in flags:
            if employee_id in id_to_row and date_col in date_to_col:
                row = id_to_row[employee_id]
                col = date_to_col[date_col]
                cell = worksheet.cell(row=row, column=col)
                cell.fill = flag_fill
                cell.font = flag_font
            else:
                print(f'Warning: Flag for ({employee_id=}, {date_col=}) not found in DataFrame.')

        # 6. Add Comments
        for (employee_id, date_col), comment_text in comments.items():
            if employee_id in id_to_row and date_col in date_to_col:
                row = id_to_row[employee_id]
                col = date_to_col[date_col]
                cell = worksheet.cell(row=row, column=col)
                n_lines = len(comment_text.splitlines())
                cell.comment = Comment(comment_text, '[rs_uy robot]', height=n_lines * 20, width=180)
            else:
                print(f'Warning: Comment for ({employee_id=}, {date_col=}) not found in DataFrame.')

        # 7. Adjust Column Widths
        for i, column_name in enumerate(df.columns, 1): # Start from 1 for col_idx
            column_letter = get_column_letter(i)
            if column_name in [id_col, name_col]:
                max_length = max(
                    df[column_name].astype(str).map(len).max(),
                    len(column_name)
                )
                worksheet.column_dimensions[column_letter].width = max_length + 4
            else:
                worksheet.column_dimensions[column_letter].width = 10

        # 8. Freeze Panes
        worksheet.freeze_panes = 'C2' # Using the cell address directly is common and clear

    return final_output_path


def run_final_sanity_checks(df: pd.DataFrame, flags: dict[str, list],
                            short_day_thresh: timedelta = timedelta(hours=2),
                            long_day_thresh: timedelta = timedelta(hours=15),
                            punch_count_thresh: int = 10) -> pd.DataFrame:
    """
    Performs final checks on the cleaned data to flag unusual patterns for a day.

    Args:
        df: The fully processed DataFrame (before helper columns are dropped).
        flags: The dictionary for recording issues.
        short_day_thresh: Any workday shorter than this duration is flagged.
        long_day_thresh: Any workday longer than this duration is flagged.
        punch_count_thresh: Any day with more than this many punches is flagged.

    Returns:
        The DataFrame, passed through without modification.
    """
    if df.empty:
        return df

    daily_summary = df.groupby(['ID', 'DATE'])

    for (employee_id, row_date), group in daily_summary:
        flag_item = {'ID': employee_id, 'DATE': row_date}

        # Check 1: Excessive punch count
        if len(group) > punch_count_thresh:
            flags.setdefault('excessive_punches', []).append(flag_item)

        # Check 2: Imbalanced day (only INs or only OUTs)
        if group['TYPE'].nunique() == 1:
            flags.setdefault('imbalanced_day', []).append(flag_item)
            continue  # Skip duration checks for imbalanced days

        # Check 3 & 4: Workday duration
        first_punch_time = group['DATETIME'].min()
        last_punch_time = group['DATETIME'].max()
        work_duration = last_punch_time - first_punch_time

        if timedelta(seconds=0) < work_duration < short_day_thresh:
            flags.setdefault('short_duration_day', []).append(flag_item)

        if work_duration > long_day_thresh:
            flags.setdefault('long_duration_day', []).append(flag_item)

    return df

def process_timesheet(df: pd.DataFrame,
                      buffer: timedelta = timedelta(minutes=15),
                      start_hour: timedelta | None = str_to_delta('07:00 AM'),
                      end_hour: timedelta | None = str_to_delta('10:00 PM'),
                      break_time: dict[str, dict] | None = None,
                      first_in_thresh: timedelta = str_to_delta('09:30 AM'),
                      last_out_thresh: timedelta = str_to_delta('02:30 PM'),
                      round_to: list[timedelta] | None = None,
                      output_filename: str = 'timesheet_summary.xlsx') -> Path | None:
    """
    Main orchestration function to process a raw timesheet DataFrame.

    This function executes the entire data cleaning and summarization pipeline
    in a specific order.

    Args:
        df: The raw DataFrame read from a CSV file.
        buffer: General time buffer for comparing punches.
        start_hour: The official start of the workday.
        end_hour: The official end of the workday.
        break_time: Dictionary defining break periods.
        first_in_thresh: Time threshold to identify a morning punch.
        last_out_thresh: Time threshold to identify an afternoon punch.
        round_to: A list of specific times to round nearby punches to.
        output_filename: The name of the output Excel file.
    """
    if break_time is None:
        break_time = {}

    if round_to is None:
        round_to = []

    # convert str time to timedelta
    for key, value in break_time.items():
        paid: bool = value['paid']
        converted_break_time: dict = {inner_key: str_to_delta(inner_value) for inner_key, inner_value in value.items()
                                      if isinstance(inner_value, str)}
        converted_break_time['paid'] = paid
        break_time[key] = converted_break_time

    # Pre-process
    df = standardize_logtype(df)
    df = sort_df(df)
    original_timestamps: dict[tuple[str, str], str] = record_timestamps(df, 'Raw Data:')
    df = add_helper_cols(df)

    flags: dict[str, list] = {}

    # Step 1: Auto-adjust Early Clock-ins to start_hour
    if start_hour is not None:
        df = adjust_first_in(df, start_hour, buffer, flags)

    # Step 2: Auto-adjust Late Clock-outs to end_hour
    if end_hour is not None:
        df = adjust_last_out(df, end_hour, buffer, flags)

    # Step 3: Verify first_in and last_out
    df = verify_first_in(df, break_time, first_in_thresh, buffer, flags)
    df = verify_last_out(df, break_time, last_out_thresh, buffer, flags)

    # Step 4: Remove duplicates of the same punch (in-in or out-out)
    df = remove_same_type_duplicate(df, buffer)

    # Step 5: Verify lunch out and lunch in
    df = adjust_break_time(df, break_time, buffer, flags)

    # Step 6: Remove duplicates of different punch (in-out or out-in)
    df = remove_diff_type_duplicate(df, buffer)

    # Step 7: In-between verified
    df = verify_in_betweens(df, flags)

    # Step 8: In-Out pair verification
    df = pair_verification(df, flags)

    # Step 9: Rounding
    df = adjust_and_round(df, round_to, buffer)

    # Step 10: Insert break time punches
    df = insert_breaks(df, break_time, flags)

    # Step 11: Flag odd groups
    df = flag_odd_groups(df, flags)

    # Step 12: Final sanity check and flag
    df = run_final_sanity_checks(df, flags)

    # Step 13: Remove helper columns
    df = df.drop(['TIMEDELTA', 'INDEX', 'VERIFIED'], axis='columns')

    # create summary
    new_timestamps: dict[tuple[str, str], str] = record_timestamps(df, 'Processed Data:')
    summary: pd.DataFrame = create_summary(df)
    comments: dict[tuple[str, str], str] = combine_comments(original_timestamps, new_timestamps, flags)
    crucial_flags: list[tuple[str, str]] = get_crucial_flags(flags)
    return create_sheet(summary, comments, crucial_flags, output_filename)

def run_default():
    # Parameters
    file_path: Path = Path('sample.csv')
    buffer: timedelta = timedelta(minutes=15)
    start_hour: timedelta = str_to_delta('07:00 AM')
    end_hour: timedelta = str_to_delta('10:00 PM')
    first_in_thresh: timedelta = str_to_delta('09:30 AM')
    last_out_thresh: timedelta = str_to_delta('02:30 PM')
    break_time = {'lunch': {'start': '12:00 PM', 'end': '01:00 PM', 'paid': False},
                  'dinner': {'start': '06:00 PM', 'end': '06:30 PM', 'paid': True}}
    round_to = ['04:00 PM', '05:00 PM', '06:00 PM']
    round_to = [str_to_delta(i) for i in round_to]
    output_filename: str = 'timesheet_summary.xlsx'

    # Process
    try:
        df, logs = read_input_file(file_path)
        for log_message in logs:
            print(log_message)
    except (ValueError, FileNotFoundError) as e:
        print(f'Error: Could not read input file. Details: {e}')
        return

    process_timesheet(df, buffer, start_hour, end_hour, break_time, first_in_thresh,
                      last_out_thresh, round_to, output_filename)
    return

if __name__ == '__main__':
    run_default()